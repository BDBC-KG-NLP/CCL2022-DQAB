import json
import torch
from transformers import AutoTokenizer, AutoModelForQuestionAnswering
from openpyxl import load_workbook


content_json_dic = {}
temp_list = []

save_out_example = []


def loadSheet(xlsx_path):
    global content_json_dic
    over_len_cnt = 0
    wb = load_workbook(xlsx_path)
    ws = wb['content1的副本']
    for r in range(1,ws.max_row):
        id = ws.cell(r,4).value
        #print(id)
        content = ws.cell(r, 3).value
        content_json_dic[id] = content

    return content_json_dic


def get_text_by_detail(data):
    ans = ""
    index = data["content-key"]
    path = data["detail"]
    start = data["location"][0]
    end = data["location"][1]
    content = eval(content_json_dic[index])
    for r in range(len(path)):
        if r == 0:
            tmp = content[path[r]]
        else:
            tmp = tmp[path[r]]
    #print(tmp)
    ans += tmp
    #print(ans)
    return ans


def load_raw_data(file_path="./temp.txt"):
    raw_datas = []
    for json_data in open(file_path, 'r', encoding='utf-8'):
        raw_data = json.loads(json_data)
        raw_datas.append(raw_data)
    return raw_datas


def write_examples(output_path, exps):
    # print("begin write ", output_path)
    with open(output_path, 'a+', encoding='utf-8') as wf:
        for exp in exps:
            json_str = json.dumps(exp, ensure_ascii=False)
            wf.writelines(json_str + '\n')
    wf.close()


def reprocess(example):
    question = example['question']
    aim = example['aim']
    answer_list = example['answer']
    type_index = 0

    key_cnt_map = {}
    for ans in answer_list:
        key = ans['content-key']
        if key in key_cnt_map.keys():
            key_cnt_map[key] += 1
        else:
            key_cnt_map[key] = 1

    key_type_map = {}
    for k,v in key_cnt_map.items():
        if v == 1:
            key_type_map[k] = "n"
        else:
            key_type_map[k] = f"y_{type_index}"
            type_index += 1

    for ans in answer_list:
        key = ans['content-key']
        r = key_cnt_map[key]
        if r == 1:
            ans["detail-type"] = "n"
            ans["correlation"] = "c"
        elif r > 1:
            ans["detail-type"] = key_type_map[key]
            ans["correlation"] = "d"

    return example


def main():
    global content_json_dic
    xlsx_path = "./content.xlsx"
    file_path = "./task2_result.txt"
    model_path = "./models/MRC_model"
    output_path = "./answer_example_para.txt"
    content_json_dic = loadSheet(xlsx_path)
    raw_datas = load_raw_data(file_path)

    tokenizer = AutoTokenizer.from_pretrained(model_path)
    model = AutoModelForQuestionAnswering.from_pretrained(model_path)
    cnt = 0
    for data in raw_datas:
        
        question = data['question']
        aim = data['aim']
        answer_list = data['answer']
        temp_out_list = []
        mrc_question = question + aim

        for ans_temp in answer_list:
            key = ans_temp['content-key']
            context = get_text_by_detail(ans_temp)

            if len(context) < 450:
                inputs = tokenizer.encode_plus(mrc_question, context, return_tensors="pt")
                outputs = model(**inputs)

                answer_start_scores = outputs[0]
                answer_end_scores = outputs[1]
                answer_start = torch.argmax(answer_start_scores)
                answer_end = torch.argmax(answer_end_scores) + 1

                result = tokenizer.convert_tokens_to_string(
                    tokenizer.convert_ids_to_tokens(inputs["input_ids"][0][answer_start:answer_end]))
                result = result.replace(" ", "")
                
                if result not in {"[CLS]", "[SEP]", ""}:
                    # print(result)
                    temp = {
                        "content-key":key,
                        'detail':ans_temp['detail'],
                        "location": [answer_start.item(), answer_end.item()]
                    }
                    temp_out_list.append(temp)
            else:
                content_start_point = 0
                answer_start_point = -1
                answer_end_point = -1
                window_len = 400
                while True:
                    content_end_point = min(content_start_point + window_len, len(context))
                    new_context = context[content_start_point:content_end_point]
                    inputs = tokenizer.encode_plus(mrc_question, new_context, return_tensors="pt")
                    outputs = model(**inputs)

                    answer_start_scores = outputs[0]
                    answer_end_scores = outputs[1]
                    answer_start = torch.argmax(answer_start_scores)
                    answer_end = torch.argmax(answer_end_scores) + 1

                    result = tokenizer.convert_tokens_to_string(
                        tokenizer.convert_ids_to_tokens(inputs["input_ids"][0][answer_start:answer_end]))
                    result = result.replace(" ", "")
                    # print(result)
                    if result not in {"[CLS]", "[SEP]", ""}:
                        if answer_start_point == -1:
                            answer_start_point = content_start_point + answer_start.item()
                        answer_end_point = content_start_point + answer_end.item()
                    if content_start_point + window_len > len(context):
                        break
                    else:
                        content_start_point += window_len // 2
                if answer_start_point == -1 and answer_end_point == -1:
                    temp = {
                        "content-key": key,
                        'detail': ans_temp['detail'],
                        "location": [content_start_point, content_end_point]
                    }
                    temp_out_list.append(temp)

        example = {
            "question":question,
            "aim": aim,
            "answer":temp_out_list
        }
        example = reprocess(example)
        print(example)
        save_out_example.append(example)
        write_examples(output_path, [example])


if __name__ == '__main__':
    main()
